# import blpapi
import re
import pandas as pd
from xbbg import blp
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment 

# Corrected code
def get_bloomberg_data(tickers, fields):
    # Get today's date
    today = datetime.today()

    # Subtract one day from today's date to get the previous working day
    # If today is a Monday (weekday 0), subtract three days to get the previous Friday
    if today.weekday() == 0:
        previous_working_day = today - timedelta(days=3)
    else:
        previous_working_day = today - timedelta(days=1)

    # Subtract one more day to get the day before the previous working day
    day_before_previous = previous_working_day - timedelta(days=1)
    # If the day before previous is a Sunday (weekday 6), subtract two more days to get the previous Friday
    if day_before_previous.weekday() == 6:
        day_before_previous = day_before_previous - timedelta(days=2)

    # Subtract seven days to get the same day from the previous week
    same_day_last_week = previous_working_day - timedelta(days=7)

    # Subtract one month from the previous working day
    one_month_ago = previous_working_day - timedelta(days=30)
    one_year_ago = previous_working_day - timedelta(days=365)

    # Format the dates in the desired format (yyyy-mm-dd)
    formatted_date_str_previous = previous_working_day.strftime('%Y-%m-%d')
    formatted_date_str_day_before = day_before_previous.strftime('%Y-%m-%d')
    formatted_date_str_last_week = same_day_last_week.strftime('%Y-%m-%d')
    formatted_date_str_one_month_ago = one_month_ago.strftime('%Y-%m-%d')
    formatted_date_str_one_year_ago = one_year_ago.strftime('%Y-%m-%d')
    # Retrieve data from Bloomberg for the previous working day
    df_bloom_previous = blp.bdh(tickers=tickers, flds=fields, start_date=formatted_date_str_previous, Per='D')

    # Retrieve data from Bloomberg for the day before the previous working day
    df_bloom_day_before = blp.bdh(tickers=tickers, flds=fields, start_date=formatted_date_str_day_before, Per='D')

    # Retrieve data from Bloomberg for the same day last week
    df_bloom_last_week = blp.bdh(tickers=tickers, flds=fields, start_date=formatted_date_str_last_week, Per='D')

    # Retrieve data from Bloomberg for one month ago from the previous working day
    df_bloom_one_month_ago = blp.bdh(tickers=tickers, flds=fields, start_date=formatted_date_str_one_month_ago, Per='D')
    df_bloom_one_year_ago = blp.bdh(tickers=tickers, flds=fields, start_date=formatted_date_str_one_year_ago, Per='D')

    # Ensure the index is in datetime format
    df_bloom_previous.index = pd.to_datetime(df_bloom_previous.index)
    df_bloom_day_before.index = pd.to_datetime(df_bloom_day_before.index)
    df_bloom_last_week.index = pd.to_datetime(df_bloom_last_week.index)
    df_bloom_one_month_ago.index = pd.to_datetime(df_bloom_one_month_ago.index)
    df_bloom_one_year_ago.index = pd.to_datetime(df_bloom_one_year_ago.index)

    # Check if the DataFrames are empty or if the indices contain NaT
    if df_bloom_previous.empty or df_bloom_previous.index.isnull().any():
        print("No data available for the previous working day.")
    else:
        # Print the results for the previous working day
        print(f"Previous working day: {previous_working_day.strftime('%d/%b/%Y')}")
        print(df_bloom_previous)
        return df_bloom_previous

    if df_bloom_day_before.empty or df_bloom_day_before.index.isnull().any():
        print("No data available for the day before the previous working day.")
    else:
        # Print the results for the day before the previous working day
        print(f"Day before the previous working day: {day_before_previous.strftime('%d/%b/%Y')}")
        print(df_bloom_day_before)

    if df_bloom_last_week.empty or df_bloom_last_week.index.isnull().any():
        print("No data available for the same day last week.")
    else:
        # Print the results for the same day last week
        print(f"Same day last week: {same_day_last_week.strftime('%d/%b/%Y')}")
        print(df_bloom_last_week)

    if df_bloom_one_month_ago.empty or df_bloom_one_month_ago.index.isnull().any():
        print("No data available for one month ago from the previous working day.")
    else:
        # Print the results for one month ago from the previous working day
        print(f"One month ago from the previous working day: {one_month_ago.strftime('%d/%b/%Y')}")
        print(df_bloom_one_month_ago)

    if df_bloom_one_year_ago.empty or df_bloom_one_year_ago.index.isnull().any():
        print("No data available for one month ago from the previous working day.")
    else:
        # Print the results for one month ago from the previous working day
        print(f"One month ago from the previous working day: {one_year_ago.strftime('%d/%b/%Y')}")
        print(df_bloom_one_year_ago)

# Define the tickers and fields
tickers = ['LQM1SPVL Index', 'LQM1RTVL Index', 'LQM1HOVL Index', 'LQM1FJVL Index', 'LQ11SPLM Index', 'LQ11FJLM Index', 'LQ11HOLM Index', 'LQ11RTLM Index', 'LQM5SP38 Index', 'LQM4FJ38 Index', 'LQM1HO38 Index', 'LQM2RD38 Index']
fields = 'PX_LAST'

# Call the function without providing a date
df_bloom_data = get_bloomberg_data(tickers, fields)
# df_bloom_previous = df_bloom_data.get('df_bloom_previous')

# wb = load_workbook("Bunker Prices.xlsx")
# ws = wb['Cover Top Ports']

# ws[f'E6'] = df_bloom_previous.iloc[:, 2] 
if 'df_bloom_previous' in df_bloom_data:
    df_bloom_previous = df_bloom_data['df_bloom_previous']

    # Verify that the DataFrame is not empty
    if not df_bloom_previous.empty:
        # Load the workbook and worksheet
        wb = load_workbook("Bunker Prices.xlsx")
        ws = wb['VLSFO']

        date_report_bbg = datetime.strftime("%d/%b/%Y")
        print(date_report_bbg)
        # ws = wb['BBG Prices']

        column = 'C'
        dates = []
        for cell in ws[column]:
            if isinstance(cell.value, str) and cell.value.startswith('=DATE('):
                match = re.match(r'=DATE\((\d+),(\d+),(\d+)\)', cell.value)
                if match:
                    year = int(match.group(1))
                    month = int(match.group(2))
                    day = int(match.group(3))
                    dates.append(datetime(year, month, day))
        # print(dates)          
        if dates:
            last_date = max(dates)
            next_date = last_date + timedelta(days=1)  
            year = next_date.year   
            last_row = ws.max_row
            print("hello")
            print(ws.max_row)
            while last_row >1 and not ws[f'{column}{last_row}'].value:
                last_row -= 1
            next_row = last_row + 1
            next_cell = ws[f'{column}{next_row}']
            next_year_cell = ws[f'A{next_row}']
            next_month_cell = ws[f'B{next_row}']
            ###################### DATE ##################
            next_cell.value = f'=DATE({date_report_bbg.year},{date_report_bbg.month},{date_report_bbg.day})' #maybe could i put the header as a date??? 
            next_cell.number_format = 'dd/mm/yyyy'
            next_cell.alignment = Alignment(horizontal='center')
            next_cell.font = Font(bold=True, color='000000')
            ###################### YEAR ##################
            next_year_cell.value = f'=YEAR(C{next_row})'
            next_year_cell.alignment = Alignment(horizontal='center')
            next_year_cell.font = Font(bold=True, color='0070C0')
            ###################### MONTH ##################
            next_month_cell.value = f'=MONTH(C{next_row})'
            next_month_cell.alignment = Alignment(horizontal='center')
            next_month_cell.font = Font(bold=True, color='0070C0')
        else:
            print("No dates found in the specified column")
        wb.close()
        print("Value assigned successfully.")
    else:
        print("The 'df_bloom_previous' DataFrame is empty.")
else:
    print("No data available for 'df_bloom_previous'.")


